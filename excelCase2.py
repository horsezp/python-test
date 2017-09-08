#!/usr/bin/python
#coding=utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

workbook_ = load_workbook('D:/Users/ma.zipeng/Desktop/BI/mapping.xlsx')

sheetnames =workbook_.get_sheet_names() #获得表单名字

print sheetnames

sheet = workbook_.get_sheet_by_name(sheetnames[0])

print sheet.cell(row=2,column=3).value

sheet2 = workbook_.get_sheet_by_name(sheetnames[2])

sheet2['C7'] = sheet['C2'].value
sheet2['C14'] = sheet['C3'].value
sheet2['C23'] = sheet['C4'].value
sheet2['C24'] = sheet['C5'].value
sheet2['C25'] = sheet['C6'].value
sheet2['C27'] = sheet['C7'].value
sheet2['C28'] = sheet['C8'].value
sheet2['C30'] = sheet['C9'].value
sheet2['C31'] = sheet['C10'].value
sheet2['C33'] = sheet['C11'].value
sheet2['C34'] = sheet['C12'].value
sheet2['C35'] = sheet['C13'].value
#sheet2['C37'] = sheet['C14'].value
sheet2['C38'] = sheet['C15'].value
sheet2['C41'] = sheet['C16'].value
#sheet2['C47'] = sheet['C17'].value
#sheet2['C49'] = sheet['C18'].value
sheet2['C63'] = sheet['C19'].value
#C48 tax
sheet2['C48'] = sheet['C17'].value - sheet['C18'].value


sheet2['D7'] = sheet['E2'].value
sheet2['D14'] = sheet['E3'].value
sheet2['D23'] = sheet['E4'].value
sheet2['D24'] = sheet['E5'].value
sheet2['D25'] = sheet['E6'].value
sheet2['D27'] = sheet['E7'].value
sheet2['D28'] = sheet['E8'].value
sheet2['D30'] = sheet['E9'].value
sheet2['D31'] = sheet['E10'].value
sheet2['D33'] = sheet['E11'].value
sheet2['D34'] = sheet['E12'].value
sheet2['D35'] = sheet['E13'].value
#sheet2['D37'] = sheet['E14'].value
sheet2['D38'] = sheet['E15'].value
sheet2['D41'] = sheet['E16'].value
#sheet2['D47'] = sheet['E17'].value
#sheet2['D49'] = sheet['E18'].value
sheet2['D63'] = sheet['E19'].value
#D48 tax
sheet2['D48'] = sheet['E17'].value - sheet['E18'].value

sheet2['E7'] = sheet['D2'].value
sheet2['E14'] = sheet['D3'].value
sheet2['E23'] = sheet['D4'].value
sheet2['E24'] = sheet['D5'].value
sheet2['E25'] = sheet['D6'].value
sheet2['E27'] = sheet['D7'].value
sheet2['E28'] = sheet['D8'].value
sheet2['E30'] = sheet['D9'].value
sheet2['E31'] = sheet['D10'].value
sheet2['E33'] = sheet['D11'].value
sheet2['E34'] = sheet['D12'].value
sheet2['E35'] = sheet['D13'].value
#sheet2['E37'] = sheet['D14'].value
sheet2['E38'] = sheet['D15'].value
sheet2['E41'] = sheet['D16'].value
#sheet2['E47'] = sheet['D17'].value
#sheet2['E49'] = sheet['D18'].value
sheet2['E63'] = sheet['D19'].value
#E48 tax
sheet2['E48'] = sheet['D17'].value - sheet['D18'].value

workbook_.save('D:/Users/ma.zipeng/Desktop/BI/mapping.xlsx');